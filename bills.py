from barcode_reader import *
from imbox import Imbox
from datetime import datetime, timedelta
import pandas as pd
import uuid
from openpyxl import Workbook

username = 'rodrigo55pereira@gmail.com'
password = open('passwords/token', 'r').read()
host = "imap.gmail.com"

download_folder = "boletos"

mail = Imbox(host, username=username, password=password, ssl=True)
messages = mail.messages(date__gt=datetime.today() - timedelta(days=30), raw="has:attachment")

path = 'no_bill'
if not os.path.isdir(path):
    os.makedirs(path)

wb = Workbook()
ws = wb.active
r = 1
ws.cell(row=1, column=1).value = "Assunto"
ws.cell(row=1, column=2).value = "Código de barras"
ws.cell(row=1, column=3).value = "Linha digitável"
ws.cell(row=1, column=4).value = "Filename"

data_atual = datetime.now()
data_atual_formatada = data_atual.strftime("%d_%m_%Y")

for (uid, message) in messages:
    for attach in message.attachments:
        att_file = attach.get('filename')

        if '.pdf' in att_file:
            print(message.subject, '-', att_file)

            unique_filename = f'{data_atual_formatada}_{str(uuid.uuid4())}'
            download_path = f'{download_folder}/{unique_filename}.pdf'
            date = pd.to_datetime(message.date)

            file = open(download_path, 'wb')
            file.write(attach.get('content').read())

            try:
                barcode = barcode_reader(download_path)
                linha_dig = linha_digitavel(barcode)
            except:
                barcode = False
            
            if not barcode:
                os.rename(download_path, os.path.join(path, unique_filename + '.pdf'))
            
            else:
               r += 1
               # Gravando no Excel
               ws.cell(row=r, column=1).value = message.subject
               ws.cell(row=r, column=2).value = barcode
               ws.cell(row=r, column=3).value = linha_dig
               ws.cell(row=r, column=3).value = unique_filename

wb.save("boletos.xlsx")
mail.logout()




